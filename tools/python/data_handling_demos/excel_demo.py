""" Simple demo of reading an Excel file in Python.

	You must install the OpenPyXL third-party library...

	NOTE: 	Rather than use a separate Python virtual env, I will use the "pip"
		command with the "--user" option when installing the package:

	python3 -m pip install --user wheel		(this is needed for the "bdist_wheel" command - see https://pypi.org/project/wheel/)
	python3 -m pip install --user openpyxl

	Using the "--user" option means that the packages will be installed under "~/" rather
	than with the system installation of Python. Here's a snippet of the files added,
	under "~/":

	./.local/lib/python3.6/site-packages/openpyxl-2.6.2.egg-info/top_level.txt
	./.local/lib/python3.6/site-packages/openpyxl-2.6.2.egg-info/SOURCES.txt
	./.local/lib/python3.6/site-packages/openpyxl-2.6.2.egg-info/dependency_links.txt
	./.local/lib/python3.6/site-packages/wheel-0.33.4.dist-info/LICENSE.txt
	./.local/lib/python3.6/site-packages/wheel-0.33.4.dist-info/METADATA
	./.local/lib/python3.6/site-packages/wheel-0.33.4.dist-info/top_level.txt
"""

from openpyxl import load_workbook

wb = load_workbook("temp_data.xlsx")
results = []
ws = wb.worksheets[0]
for row in ws.iter_rows():
	results.append([cell.value for cell in row])

print(results)
